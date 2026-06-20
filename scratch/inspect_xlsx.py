import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r"c:\Users\panga\OneDrive\Desktop\Seva\2026Yatra\src\components\Accounts\Accounts.xlsx", data_only=True)
sheet = wb.active

headers = [cell.value for cell in sheet[2]]
for i, h in enumerate(headers):
    print(f"Col {i+1} ({get_column_letter(i+1)}): {h}")

# Let's inspect first few rows of data starting from row 3
print("\nFirst 10 data rows:")
for r_idx in range(3, 13):
    row_vals = [cell.value for cell in sheet[r_idx]]
    if any(row_vals):
        print(f"Row {r_idx}: {[row_vals[0], row_vals[1], row_vals[2], row_vals[19], row_vals[20], row_vals[21]]}")
